"""Local Excel (.xlsx) export for RecContinue (patient session + clinic dashboard).

Mirrors the privacy stance of packet_export.py: everything here reads from the
local session dict / local vault and writes a single local .xlsx file. Nothing
is transmitted anywhere.
"""
from __future__ import annotations

import pathlib
from datetime import datetime, timezone
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="14253D", end_color="14253D", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")


def _autosize(ws, widths: list[int]) -> None:
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def _write_header(ws, row: int, headers: list[str]) -> None:
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def _kv_rows(ws, start_row: int, pairs: list[tuple[str, Any]]) -> int:
    row = start_row
    for label, value in pairs:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=row, column=2, value=value if value not in (None, "") else "Not recorded")
        cell.alignment = WRAP
        row += 1
    return row


def build_session_workbook(session: dict[str, Any], output_dir: str | pathlib.Path) -> str:
    """Export one session (metrics + report) as a clinic-friendly .xlsx workbook."""
    patient = session.get("patient") or {}
    metrics = session.get("movement_metrics") or {}
    report = session.get("gemma_report") or {}
    context = session.get("patient_context") or {}

    wb = Workbook()

    summary = wb.active
    summary.title = "Summary"
    summary.append(["RecContinue Session Summary"])
    summary["A1"].font = Font(size=14, bold=True, color="14253D")
    summary.append([])
    row = _kv_rows(summary, 3, [
        ("Session ID", session.get("session_id")),
        ("Patient ID", patient.get("patient_id")),
        ("Task", patient.get("task")),
        ("Review status", session.get("review_status")),
        ("Data status", patient.get("data_status")),
        ("Exported at (UTC)", datetime.now(timezone.utc).isoformat()),
    ])
    summary.append([])
    summary.cell(row=row + 1, column=1, value="Plain-language recap").font = Font(bold=True)
    recap_cell = summary.cell(row=row + 2, column=1, value=report.get("patient_friendly_recap", "Not available."))
    recap_cell.alignment = WRAP
    summary.merge_cells(start_row=row + 2, start_column=1, end_row=row + 2, end_column=4)
    _autosize(summary, [26, 60, 16, 16])

    m_sheet = wb.create_sheet("Movement Metrics")
    _write_header(m_sheet, 1, ["Measurement", "Value"])
    labels = {
        "observation_module": "Observed module",
        "selected_arm": "Selected arm",
        "detected_frame_count": "Detected frames",
        "observed_2d_elbow_angle_min_degrees": "Elbow angle - min (deg)",
        "observed_2d_elbow_angle_max_degrees": "Elbow angle - max (deg)",
        "observed_2d_elbow_angle_change_degrees": "Elbow angle - change (deg)",
        "repetition_count": "Repetition count",
        "session_duration_seconds": "Session duration (s)",
        "max_hand_height_relative_to_shoulder": "Max hand height vs. shoulder",
        "observed_head_turn_proxy_range": "Head turn range (proxy)",
        "observed_head_position_variation": "Head position variation",
        "palm_opening_ratio_min": "Palm opening ratio - min",
        "palm_opening_ratio_max": "Palm opening ratio - max",
        "palm_opening_ratio_change": "Palm opening ratio - change",
    }
    r = 2
    for key, label in labels.items():
        if key in metrics:
            m_sheet.cell(row=r, column=1, value=label)
            m_sheet.cell(row=r, column=2, value=metrics[key])
            r += 1
    conf = metrics.get("landmark_confidence") or {}
    if conf:
        m_sheet.cell(row=r, column=1, value="Landmark confidence - low confidence flag")
        m_sheet.cell(row=r, column=2, value=bool(conf.get("low_confidence")))
        r += 1
    _autosize(m_sheet, [38, 24])

    ctx_sheet = wb.create_sheet("Patient Context")
    _write_header(ctx_sheet, 1, ["Field", "Patient-reported information"])
    r = 2
    for key, value in context.items():
        ctx_sheet.cell(row=r, column=1, value=key.replace("_", " ").title())
        cell = ctx_sheet.cell(row=r, column=2, value=value)
        cell.alignment = WRAP
        r += 1
    _autosize(ctx_sheet, [26, 70])

    rep_sheet = wb.create_sheet("Clinical Report")
    _write_header(rep_sheet, 1, ["Section", "Content"])
    sections = [
        ("Report status", report.get("report_status")),
        ("Session summary", report.get("session_summary")),
        ("Objective observations", "\n".join(f"- {x}" for x in report.get("objective_observations", []))),
        ("Patient-reported information", "\n".join(f"- {x}" for x in report.get("patient_reported_information", []))),
        ("Person factors", "\n".join(f"- {x}" for x in report.get("person_factors", []))),
        ("Environment factors", "\n".join(f"- {x}" for x in report.get("environment_factors", []))),
        ("Occupation factors", "\n".join(f"- {x}" for x in report.get("occupation_factors", []))),
        ("Missing information", "\n".join(f"- {x}" for x in report.get("missing_information", []))),
        ("Clinician follow-up questions", "\n".join(f"- {x}" for x in report.get("clinician_follow_up_questions", []))),
        ("Limitations", "\n".join(f"- {x}" for x in report.get("limitations", []))),
        ("Safety notice", report.get("safety_notice")),
    ]
    r = 2
    for label, value in sections:
        rep_sheet.cell(row=r, column=1, value=label).font = Font(bold=True)
        cell = rep_sheet.cell(row=r, column=2, value=value or "Not available.")
        cell.alignment = WRAP
        rep_sheet.row_dimensions[r].height = 45
        r += 1
    _autosize(rep_sheet, [30, 90])

    output_dir = pathlib.Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    patient_id = patient.get("patient_id") or "session"
    output_path = output_dir / f"{patient_id}-session.xlsx"
    wb.save(output_path)
    return str(output_path)


def build_vault_workbook(sessions: list[dict[str, Any]], output_dir: str | pathlib.Path) -> str:
    """Export the full local session vault as one clinic tracking .xlsx workbook.

    `sessions` is the list of full session dicts (as loaded via storage.load_session
    for each entry in storage.list_sessions()), so the sheet can include review
    status alongside the key measurements clinic staff track between visits.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Clinic Tracking"
    headers = [
        "Session ID", "Patient ID", "Task", "Saved at (UTC)", "Review status",
        "Observed module", "Detected frames", "Repetition count",
        "Elbow angle change (deg)", "Low landmark confidence", "Safety check passed",
    ]
    _write_header(ws, 1, headers)
    for r, session in enumerate(sessions, start=2):
        patient = session.get("patient") or {}
        metrics = session.get("movement_metrics") or {}
        safety = session.get("safety_check") or {}
        conf = metrics.get("landmark_confidence") or {}
        ws.cell(row=r, column=1, value=session.get("session_id"))
        ws.cell(row=r, column=2, value=session.get("patient_id") or patient.get("patient_id"))
        ws.cell(row=r, column=3, value=patient.get("task"))
        ws.cell(row=r, column=4, value=session.get("saved_at"))
        ws.cell(row=r, column=5, value=session.get("review_status"))
        ws.cell(row=r, column=6, value=metrics.get("observation_module"))
        ws.cell(row=r, column=7, value=metrics.get("detected_frame_count"))
        ws.cell(row=r, column=8, value=metrics.get("repetition_count"))
        ws.cell(row=r, column=9, value=metrics.get("observed_2d_elbow_angle_change_degrees"))
        ws.cell(row=r, column=10, value=bool(conf.get("low_confidence")))
        ws.cell(row=r, column=11, value=safety.get("passed"))
    _autosize(ws, [24, 16, 34, 24, 34, 16, 14, 14, 18, 18, 16])

    output_dir = pathlib.Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / "reccontinue-clinic-tracking.xlsx"
    wb.save(output_path)
    return str(output_path)
